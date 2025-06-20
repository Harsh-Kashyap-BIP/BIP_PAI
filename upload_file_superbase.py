import asyncio

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from io import BytesIO
import uuid
import httpx
import os
from dotenv import load_dotenv
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SERVICE_ROLE")
BUCKET_NAME = "exports"

async def upload_df_to_supabase_async(df: pd.DataFrame, file_prefix: str = "report") -> str:
    # Step 1: Save dataframe to Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    output.seek(0)

    # Step 2: Beautify the Excel file
    wb = load_workbook(output)
    ws = wb.active

    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 5

    beautified = BytesIO()
    wb.save(beautified)
    beautified.seek(0)

    # Step 3: Generate filename
    filename = f"{file_prefix}_{uuid.uuid4()}.xlsx"
    file_path = f"{BUCKET_NAME}/{filename}"

    # Step 4: Upload to Supabase Storage using HTTPX
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

    async with httpx.AsyncClient() as client:
        response = await client.post(
            f"{SUPABASE_URL}/storage/v1/object/{file_path}",
            headers=headers,
            content=beautified.read()
        )
        response.raise_for_status()  # Raise error if upload fails

    # Step 5: Generate public URL
    public_url = f"{SUPABASE_URL}/storage/v1/object/public/{file_path}"
    return public_url
